"""P1-T5: Sheets/CSV parsing — normalized rows for app_* tables.

Pure parsing: no formula execution ever. XLSX is read with openpyxl
``data_only=False`` so formula cells yield their literal formula string,
never a computed value. CSV is parsed with the stdlib csv module.
"""

import csv
import hashlib
from pathlib import Path
from typing import Any

from src.app_logging import get_logger

logger = get_logger()

MISSING_OPENPYXL_MSG = (
    "Error: .xlsx parsing requires the 'openpyxl' package "
    "(install with: uv add openpyxl); .csv files work without it."
)


def _infer_type(values: list[Any]) -> str:
    """Infer a column SQL type from its non-empty values.

    CSV values arrive as strings — numeric-looking strings infer as
    numeric types so _coerce normalizes them on insert.
    """
    non_empty = [v for v in values if v not in (None, "")]
    if not non_empty:
        return "TEXT"
    if all(isinstance(v, bool) for v in non_empty):
        return "BOOLEAN"
    if all(isinstance(v, int) and not isinstance(v, bool) for v in non_empty):
        return "INTEGER"
    try:
        [float(v) for v in non_empty]
        return "REAL"
    except (ValueError, TypeError):
        pass
    lowered = [str(v).strip().lower() for v in non_empty]
    if all(v in ("true", "false", "yes", "no", "1", "0") for v in lowered):
        return "BOOLEAN"
    return "TEXT"


def _coerce(value: Any, sql_type: str) -> Any:
    if value in (None, ""):
        return None
    try:
        if sql_type == "INTEGER":
            return int(value)
        if sql_type == "REAL":
            return float(value)
        if sql_type == "BOOLEAN" and isinstance(value, str):
            return value.strip().lower() in ("true", "yes", "1")
    except (ValueError, TypeError):
        pass
    return str(value) if not isinstance(value, (int, float, bool)) else value


class SheetParseError(Exception):
    """Raised when a workbook/csv cannot be parsed."""


def parse_csv(path: Path) -> dict[str, list[dict[str, Any]]]:
    """Parse a CSV into {table_name: rows} — table name from the file stem."""
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        if reader.fieldnames is None:
            raise SheetParseError(f"No header row found in {path.name}")
        rows: list[dict[str, Any]] = [dict(r) for r in reader]
    return {path.stem: rows}


def parse_xlsx(path: Path) -> dict[str, list[dict[str, Any]]]:
    """Parse an XLSX workbook into {sheet_name: rows} (values only, no
    formula evaluation)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SheetParseError(MISSING_OPENPYXL_MSG)
    try:
        # data_only=False: formula cells return their literal "=..." string.
        # openpyxl never executes formulas either way; this keeps raw values.
        wb = load_workbook(path, data_only=False, read_only=True)
    except Exception as exc:
        raise SheetParseError(f"Error: cannot read workbook {path.name}: {exc}")
    sheets: dict[str, list[dict[str, Any]]] = {}
    try:
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            header = next(it, None)
            if header is None:
                continue
            columns = [str(h) if h is not None else f"col{i}" for i, h in enumerate(header)]
            rows = []
            for values in it:
                if all(v in (None, "") for v in values):
                    continue
                rows.append({c: v for c, v in zip(columns, values) if c})
            if rows:
                sheets[ws.title] = rows
    finally:
        wb.close()
    return sheets


def parse_sheets(path: Path) -> dict[str, list[dict[str, Any]]]:
    """Parse any supported sheet file into {sheet_name: normalized rows}."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return parse_csv(path)
    if suffix in (".xlsx", ".xlsm"):
        return parse_xlsx(path)
    raise SheetParseError(f"Error: unsupported sheet format '{path.suffix}'")


def source_key(path: Path) -> str:
    """Stable provenance key for a parsed file (content hash + path) — used
    for duplicate-import upsert (rows from the same source are replaced)."""
    digest = hashlib.sha256(path.read_bytes()).hexdigest()[:16]
    return f"{path.name}:{digest}"


def rows_to_schema(rows: list[dict[str, Any]]) -> dict[str, str]:
    """Infer {column: sql_type} from normalized rows, including the
    provenance column ``_source_key``."""
    columns = list(rows[0].keys())
    schema = {
        col: _infer_type([r.get(col) for r in rows]) for col in columns
    }
    schema["_source_key"] = "TEXT"
    return schema


def normalize(rows: list[dict[str, Any]], sql_types: dict[str, str]) -> list[dict[str, Any]]:
    """Coerce row values to the inferred SQL types."""
    out = []
    for row in rows:
        out.append({c: _coerce(v, sql_types.get(c, "TEXT")) for c, v in row.items()})
    return out
